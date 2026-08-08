import json
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

SCRIPT = Path(__file__).resolve().parents[1] / "scripts" / "generate_field_excel.py"

FIELDS = json.dumps([
    {
        "english": "EXAM_NAME", "chinese": "考试名称", "type": "VARCHAR",
        "length": 100, "precision": None,
        "is_pk": False, "is_dist_key": False, "is_part_key": False,
        "nullable": "是", "default": None, "remark": "考试名称"
    },
    {
        "english": "RANK", "chinese": "排名", "type": "INT",
        "length": None, "precision": None,
        "is_pk": True, "is_dist_key": True, "is_part_key": True,
        "nullable": "否", "default": None, "remark": "排名（1-10）"
    },
])

HEADERS = ["属性", "英文名", "中文名", "数据类型", "长度", "精度",
           "主键", "分布键", "分区键", "是否允许空值", "默认值", "备注"]


def run_script(tmp_path, fields=FIELDS, table="ABCD_EXAM_TOP10_BY_SUBJECT"):
    out = tmp_path / f"目标表字段定义_{table}.xlsx"
    result = subprocess.run(
        [sys.executable, str(SCRIPT), "--output", str(out),
         "--table", table, "--fields", fields],
        capture_output=True, text=True,
    )
    return result, out


def test_creates_xlsx_with_12_headers(tmp_path):
    result, out = run_script(tmp_path)
    assert result.returncode == 0, result.stderr
    assert out.exists()
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    assert [c.value for c in ws[1]] == HEADERS


def test_field_rows_fill_columns(tmp_path):
    result, out = run_script(tmp_path)
    assert result.returncode == 0, result.stderr
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 2
    assert rows[0][1] == "EXAM_NAME"   # 英文名
    assert rows[0][2] == "考试名称"      # 中文名
    assert rows[0][3] == "VARCHAR"      # 数据类型
    assert rows[0][4] == 100            # 长度
    assert rows[0][5] is None           # 精度


def test_pk_row_sets_keys_and_nullable(tmp_path):
    result, out = run_script(tmp_path)
    assert result.returncode == 0, result.stderr
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    pk_row = [r for r in rows if r[1] == "RANK"][0]
    assert pk_row[6] == "是"   # 主键
    assert pk_row[7] == "是"   # 分布键
    assert pk_row[8] == "是"   # 分区键
    assert pk_row[9] == "否"   # 是否允许空值


def test_attribute_column_defaults_empty(tmp_path):
    result, out = run_script(tmp_path)
    assert result.returncode == 0, result.stderr
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert all(r[0] is None for r in rows)  # 属性列默认不填


def test_attribute_column_uses_passed_value(tmp_path):
    fields = json.loads(FIELDS)
    fields[0]["attribute"] = "目标表"
    result, out = run_script(tmp_path, fields=json.dumps(fields))
    assert result.returncode == 0, result.stderr
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert rows[0][0] == "目标表"


def test_invalid_type_aborts(tmp_path):
    bad = json.dumps([{
        "english": "BAD", "chinese": "非法类型", "type": "JSON",
        "length": None, "precision": None,
        "is_pk": False, "is_dist_key": False, "is_part_key": False,
        "nullable": "是", "default": None, "remark": ""
    }])
    result, _ = run_script(tmp_path, fields=bad)
    assert result.returncode != 0
    assert "类型" in result.stderr


def test_empty_fields_warns(tmp_path):
    result, _ = run_script(tmp_path, fields="[]")
    assert result.returncode == 0
    assert "警告" in result.stderr or "没有字段" in result.stderr