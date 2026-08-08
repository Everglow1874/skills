#!/usr/bin/env python3
"""把目标表字段定义生成 12 列 .xlsx，供平台直接导入。"""

import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

WHITELIST_TYPES = {
    "VARCHAR", "CHAR", "TINYINT", "SMALLINT", "INTEGER", "BIGINT",
    "NUMERIC", "DECIMAL", "DATE", "TIME", "TIMESTAMP",
    "TIMESTAMP WITHOUT TIME ZONE", "BOOLEAN",
}

HEADERS = ["属性", "英文名", "中文名", "数据类型", "长度", "精度",
           "主键", "分布键", "分区键", "是否允许空值", "默认值", "备注"]

YES = "是"
NO = "否"


def sanitize_filename(name: str) -> str:
    safe = re.sub(r'[\\/:*?"<>|]', "_", name)
    return safe[:120]


def validate_fields(fields):
    if not isinstance(fields, list) or not fields:
        print("警告:没有字段定义，将只输出表头。", file=sys.stderr)
        return
    for f in fields:
        ftype = str(f.get("type", "")).upper()
        if ftype not in WHITELIST_TYPES:
            raise ValueError(f"非法类型:{ftype}，不在白名单 {sorted(WHITELIST_TYPES)} 内")


def write_excel(output: Path, table: str, fields):
    validate_fields(fields)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = table[:31]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for f in fields:
        is_pk = bool(f.get("is_pk", False))
        pk_val = YES if is_pk else NO
        ws.append([
            f.get("attribute"),               # 属性（默认不填）
            f.get("english"),
            f.get("chinese"),
            str(f.get("type", "")).upper(),
            f.get("length"),
            f.get("precision"),
            pk_val,                           # 主键
            YES if f.get("is_dist_key") else NO,  # 分布键
            YES if f.get("is_part_key") else NO,  # 分区键
            NO if is_pk else "是",             # 是否允许空值
            f.get("default"),
            f.get("remark"),
        ])

    wb.save(output)


def main():
    parser = argparse.ArgumentParser(description="生成目标表字段定义 Excel")
    parser.add_argument("--output", required=True, help="输出 .xlsx 路径")
    parser.add_argument("--table", required=True, help="目标表名（含前缀全名）")
    parser.add_argument("--fields", required=True, help="JSON 数组格式的字段定义")
    args = parser.parse_args()

    try:
        fields = json.loads(args.fields)
    except json.JSONDecodeError as e:
        print(f"错误:fields 参数不是合法 JSON:{e}", file=sys.stderr)
        sys.exit(1)

    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    safe_table = sanitize_filename(args.table)
    safe_output = output.parent / (sanitize_filename(output.stem) + ".xlsx")

    try:
        write_excel(safe_output, safe_table, fields)
    except ValueError as e:
        print(f"错误:{e}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"错误:生成 Excel 失败:{e}", file=sys.stderr)
        sys.exit(1)

    print(f"已生成:{safe_output}")


if __name__ == "__main__":
    main()